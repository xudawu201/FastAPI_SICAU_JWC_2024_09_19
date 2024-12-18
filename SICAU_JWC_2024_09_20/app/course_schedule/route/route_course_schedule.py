'''
Author: xudawu
Date: 2024-11-12 10:49:26
LastEditors: xudawu
LastEditTime: 2024-12-18 17:46:47
'''

import time
import copy
import urllib.parse

# fastapi相关
from fastapi import APIRouter, Request,WebSocket
import asyncio

# 下载文件相关
import fastapi
import io
import openpyxl
import urllib
import openpyxl.styles

# # 引入模板模块
from template import TemplatesJinja2CourseSchedule
from app.course_schedule.service import service_course_schedule
from app.course_schedule.service import service_initialize_course_schedule
from app.course_schedule.service import service_course_schedule_get_semester

from app import public_function

from typing import Dict

# 全局超参数
router = APIRouter()

# 维护会话状态
session_states: Dict[str, Dict] = {}

# 初始化每代排课信息
all_generation_info_dict={}
# 每代基因最大变异次数
max_mutate_count_int = 50
# 定义周、天和时间段
week_range = range(1, 17)
day_range = [1, 2, 3, 4, 5,6]
slot_time_range =[1,2,3,4,5]
# 指定学期
semester_str ='2024-2025-1'
# global semester_str = '2024-2025-1'

# 异步缓冲队列
queue_size_int = 100
MessageQueue = asyncio.Queue(maxsize=queue_size_int)

# 渲染前端页面的路由
@router.get("/course_schedule")
async def course_schedule(request: Request):
    return TemplatesJinja2CourseSchedule.TemplateResponse("course_schedule.html", {"request": request})

# 获取筛选框数据
@router.post("/load_semester")
async def load_semester(request: Request):
    # 示例数据
    # semester_dict = {
    #     "semester_list": ["2020-2021-1", "2020-2021-2"],
    #     "message": []
    # }

    # print('开始获取数据')
    
    # 获取所有数据
    semester_dict = service_course_schedule_get_semester.initialize_semester()
   
    # 定义排序的关键函数
    def sort_key(semester_list):
        year, _, term = semester_list.split('-')
        return (int(year), int(term))

    # 使用 sorted 函数进行排序,降序排序
    sorted_semester_list = sorted(semester_dict.get('semester_list'), key=sort_key,reverse=True)

    # 更新排序后的数据
    semester_dict['semester_list'] = sorted_semester_list

    # 声明全局变量
    global semester_str 
    # 更新学期为最新的一个
    semester_str = semester_dict['semester_list'][0]

    # print('获取数据完成')

    # 返回响应
    return semester_dict

# 更新学期数据
@router.post("/update_semester")
async def update_semester(request: Request):
    # 示例数据
    # semester_dict = {
    #     "semester_list": ["2020-2020-1", "2020-2020-2"],
    #     "message": []
    # }

    # 获取请求体中的 JSON 数据
    request_data = await request.json()
    request_semester_str = request_data.get("semester_str")
    
    # 声明全局变量
    global semester_str 
    semester_str = request_semester_str
    # 更新全局学期数据

async def add_to_queue(message: Dict):
    """向队列中添加消息，如果队列已满，移除最早的消息"""
    if MessageQueue.full():
        await MessageQueue.get()  # 丢弃最早的消息
    await MessageQueue.put(message)

# 开始快速排课
@router.websocket("/ws_course_schedule_quick")
async def ws_course_schedule_quick(websocket: WebSocket):
    # 声明全局变量
    global semester_str
  
    # 上课时间初始化
    print('初始上课时间类列表开始')
    start_time = time.time()
    # 初始化上课时间列表
    TimeSlot_list = service_initialize_course_schedule.initialize_time(week_range,day_range,slot_time_range)
    time_used = public_function.get_time_used(start_time)
    print('初始上课时间列表完成,用时:', time_used)
    print('上课时间数量:',len(TimeSlot_list))

    # 初始化教室类列表
    print('初始化教室类列表开始')
    start_time = time.time()
    Room_list = service_initialize_course_schedule.initialize_room()
    time_used = public_function.get_time_used(start_time)
    print('初始化教室类列表完成,用时:', time_used)
    print('教室数量:',len(Room_list))
    
    # 初始化教师和课程类列表
    print('初始化教师和课程类列表开始')
    start_time = time.time()
    Teacher_list,Course_list = service_initialize_course_schedule.initialize_teacher_course(semester_str)
    time_used = public_function.get_time_used(start_time)
    print('初始化教师和课程类列表完成,用时:', time_used)
    print('课程人数:',len(Course_list))
    print('教师人数:',len(Teacher_list))

    # 设置教师上课限制时间
    print('设置教师上课限制时间开始')
    start_time = time.time()
    Teacher_list = service_initialize_course_schedule.set_teacher_unavailable_timeslot(Teacher_list,semester_str)
    time_used = public_function.get_time_used(start_time)
    print('设置教师上课限制时间完成,用时:', time_used)

    # 初始化学生类列表
    print('初始化学生类列表开始')
    start_time = time.time()
    student_course_dict,Student_list = service_initialize_course_schedule.initialize_student(semester_str)
    time_used = public_function.get_time_used(start_time)
    print('初始化学生类列表完成,用时:', time_used)
    print('学生人数:',len(student_course_dict))

    # 设置学生选课
    print('设置学生选课列表开始')
    start_time = time.time()
    Student_list = service_initialize_course_schedule.set_student_enroll_course(Student_list,Course_list,student_course_dict)
    time_used = public_function.get_time_used(start_time)
    print('设置学生选课列表完成,用时:', time_used)
    
    # 种群大小和适应度阈值
    population_size_int = 100

    print('排课种群初始化开始')
    start_time = time.time()
    # 获得初始化的种群
    PopulationList = service_course_schedule.initialize_population(Course_list, TimeSlot_list, Room_list, Student_list, population_size_int)
    if PopulationList == True:
        error_str = '没有可排课的教室,中断排课'
        # await websocket.send_json({'message': error_str})
        await websocket.close()
        print(error_str)
        return
    time_used = public_function.get_time_used(start_time)
    print('排课种群初始化完成,用时:', time_used)

    print('获取最优个体开始')
    start_time = time.time()
    # 获取最优个体
    BestSchedule = max(PopulationList, key=lambda x: x.fitness())
    # # 批量获取最优个体2、4讲的实验课
    # SetedSchedule_dict = BestSchedule.batch_initialize_schedule_experiment_inherit(BestSchedule.ScheduleIndex_dict)
    # # 赋予处理后的实验课基因
    # BestSchedule.ScheduleIndex_dict = SetedSchedule_dict

    time_used = public_function.get_time_used(start_time)
    print('获取最优个体完成,用时:', time_used)

    # 初始化当前最高适应度
    cur_max_fitness_float=BestSchedule.fitness()

    # 初始化进化次数
    generation_int = 0
    # 等待接受WebSocket连接
    await websocket.accept()
    # print('排课种群进化开始')
    while True:
        # 加锁以保证异步执行时，每次循环内的数据是隔离的
        # async with AsyncioLock:
        # 找到进化方向
        process_message = '尝试找到进化方向开始'
        print(process_message)
        # 添加消息到队列
        await add_to_queue({
            "process_message": process_message
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)
            
        start_time = time.time()
        while True:
            # 执行深拷贝
            CurSchedule = copy.deepcopy(BestSchedule)
            # 执行变异
            CurSchedule.mutate()
            # 执行变异2
            CurSchedule.mutate()
            # 执行变异3
            # CurSchedule.mutate()

            # 获得最优个体的适应度
            # best_fitness_float = CurSchedule.fitness()
            best_fitness_float = await asyncio.to_thread(CurSchedule.fitness)
            # time_used = public_function.get_time_used(start_time)
            # print('计算当前适应度完成,用时:', time_used)
            process_message = f'当前适应度:{best_fitness_float}'
            print(process_message)
            # 添加消息到队列
            await add_to_queue({
                "process_message": process_message
            })

            # 取出队列中的消息并发送
            while not MessageQueue.empty():
                message = await MessageQueue.get()
                await websocket.send_json(message)

            # 如果适应度降低则是不好的变异方向,则放弃,重新变异
            if best_fitness_float<=cur_max_fitness_float:
                continue
            else:
                # 适应度增加,则跳出循环
                break
        time_used = public_function.get_time_used(start_time)
        process_message = f'尝试找到进化方向完成,用时:{time_used}'
        print(process_message)
        # 添加消息到队列
        await add_to_queue({
            "process_message": process_message
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)

        # 变异方向是增加适应度的,则继续进化,基因继承和适应度继承
        BestSchedule = CurSchedule
        cur_max_fitness_float = best_fitness_float

        process_message = '处理种群信息可视化开始'
        print(process_message)
        # 添加消息到队列
        await add_to_queue({
            "process_message": process_message
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)
        start_time = time.time()
        # 打印当前代的最优适应度
        generation_int += 1
        # print(f"Generation {generation_int}: Best Fitness = {best_fitness_float}")
        # 获得当前代的排课信息
        # current_generation_info_dict = await asyncio.to_thread(CurSchedule.display)
        current_generation_info_dict = CurSchedule.display()
        # 获得已排课的课程数
        total_assigned_course_number = len(current_generation_info_dict['room_assigned_course_schedule_dict'].keys())
        # 字典添加代数和适应度
        current_generation_info_dict['generation_int']=generation_int
        # print('当前代数:',current_generation_info_dict['generation_int'])
        current_generation_info_dict['fitness_float']=best_fitness_float
        current_generation_info_dict['total_assigned_course_number']=total_assigned_course_number

        time_used = public_function.get_time_used(start_time)
        process_message = f'处理种群信息可视化完成,用时:{time_used}'
        print(process_message)
        # 添加消息到队列
        await add_to_queue({
            "process_message": process_message
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)

        # 存储全局课程排课信息
        all_generation_info_dict[generation_int] = current_generation_info_dict

        # 将内容发送给前端实时显示
        # await websocket.send_json({
        #     "generation": generation_int,"best_fitness":best_fitness_float,
        #     'current_generation_info_dict':current_generation_info_dict
        #     })
        # 添加消息到队列
        await add_to_queue({
            "generation": generation_int,"best_fitness":best_fitness_float,
            'current_generation_info_dict':current_generation_info_dict
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)

# 开始更优排课
@router.websocket("/ws_course_schedule_better")
async def ws_course_schedule_better(websocket: WebSocket):
    # 声明全局变量
    global semester_str
  
    # 上课时间初始化
    print('初始上课时间类列表开始')
    start_time = time.time()
    # 初始化上课时间列表
    TimeSlot_list = service_initialize_course_schedule.initialize_time(week_range,day_range,slot_time_range)
    time_used = public_function.get_time_used(start_time)
    print('初始上课时间列表完成,用时:', time_used)
    print('上课时间数量:',len(TimeSlot_list))

    # 初始化教室类列表
    print('初始化教室类列表开始')
    start_time = time.time()
    Room_list = service_initialize_course_schedule.initialize_room()
    time_used = public_function.get_time_used(start_time)
    print('初始化教室类列表完成,用时:', time_used)
    print('教室数量:',len(Room_list))
    
    # 初始化教师和课程类列表
    print('初始化教师和课程类列表开始')
    start_time = time.time()
    Teacher_list,Course_list = service_initialize_course_schedule.initialize_teacher_course(semester_str)
    time_used = public_function.get_time_used(start_time)
    print('初始化教师和课程类列表完成,用时:', time_used)
    print('课程人数:',len(Course_list))
    print('教师人数:',len(Teacher_list))

    # 设置教师上课限制时间
    print('设置教师上课限制时间开始')
    start_time = time.time()
    Teacher_list = service_initialize_course_schedule.set_teacher_unavailable_timeslot(Teacher_list,semester_str)
    time_used = public_function.get_time_used(start_time)
    print('设置教师上课限制时间完成,用时:', time_used)

    # 初始化学生类列表
    print('初始化学生类列表开始')
    start_time = time.time()
    student_course_dict,Student_list = service_initialize_course_schedule.initialize_student(semester_str)
    time_used = public_function.get_time_used(start_time)
    print('初始化学生类列表完成,用时:', time_used)
    print('学生人数:',len(student_course_dict))

    # 设置学生选课
    print('设置学生选课列表开始')
    start_time = time.time()
    Student_list = service_initialize_course_schedule.set_student_enroll_course(Student_list,Course_list,student_course_dict)
    time_used = public_function.get_time_used(start_time)
    print('设置学生选课列表完成,用时:', time_used)
    
    # 种群大小和适应度阈值
    population_size_int = 100

    print('排课种群初始化开始')
    start_time = time.time()
    # 获得初始化的种群
    PopulationList = service_course_schedule.initialize_population(Course_list, TimeSlot_list, Room_list, Student_list, population_size_int)
    if PopulationList == True:
        error_str = '没有可排课的教室,中断排课'
        # await websocket.send_json({'message': error_str})
        await websocket.close()
        print(error_str)
        return
    time_used = public_function.get_time_used(start_time)
    print('排课种群初始化完成,用时:', time_used)

    print('获取最优个体开始')
    start_time = time.time()
    # 获取最优个体
    BestSchedule = max(PopulationList, key=lambda x: x.fitness())
    # 批量获取最优个体2、4讲的实验课
    # SetedSchedule_dict = BestSchedule.batch_initialize_schedule_experiment_inherit(BestSchedule.ScheduleIndex_dict)
    # # 赋予处理后的实验课基因
    # BestSchedule.ScheduleIndex_dict = SetedSchedule_dict

    time_used = public_function.get_time_used(start_time)
    print('获取最优个体完成,用时:', time_used)

    # 初始化当前最高适应度
    cur_max_fitness_float=BestSchedule.fitness()

    # 初始化进化次数
    generation_int = 0
    # 初始化一个锁
    # AsyncioLock = asyncio.Lock()
    # 连接前端进行实时通信
    await websocket.accept()
    

    # print('排课种群进化开始')
    while True:
        
        # 发送心跳包
        # await websocket.send_text('ping')
        # 执行深拷贝
        CurSchedule = copy.deepcopy(BestSchedule)
        # 找到进化方向
        process_message = '尝试找到进化方向开始'
        print(process_message)
        # await websocket.send_json({
        #     "process_message": process_message,
        #     })
        # 初始化变异次数
        mutate_count_int = 0
        while True:
            # 接收前端消息
            # keyword = await websocket.receive_text()
            # print(keyword)
            # 发送心跳包
            # await websocket.send_text('ping')

            # 探索变异次数大于阈值,认为此基因的变异方向不合理,退回起点重新寻找
            if mutate_count_int>max_mutate_count_int:
                # 执行深拷贝,重新赋予起始基因
                CurSchedule = copy.deepcopy(BestSchedule)
                # 异步深拷贝 BestSchedule
                # CurSchedule = await asyncio.to_thread(copy.deepcopy, BestSchedule)
                # 重置变异次数
                mutate_count_int = 0
                process_message = '变异方向不合理,退回起点重新寻找'
                print(process_message)
                # 添加消息到队列
                await add_to_queue({
                    "process_message": process_message
                })

                # 取出队列中的消息并发送
                while not MessageQueue.empty():
                    message = await MessageQueue.get()
                    await websocket.send_json(message)
                # await websocket.send_json({
                #     "process_message": process_message,
                #     })
            # 执行变异
            CurSchedule.mutate()
            # 变异次数增加
            mutate_count_int+=1

            # 获得最优个体的适应度
            # best_fitness_float = CurSchedule.fitness()
            best_fitness_float = await asyncio.to_thread(CurSchedule.fitness)
            # time_used = public_function.get_time_used(start_time)
            # print('计算当前适应度完成,用时:', time_used)
            process_message = f'{mutate_count_int}/{max_mutate_count_int},当前适应度:{best_fitness_float}'
            print(process_message)
            # await websocket.send_json({
            #     "process_message": process_message,
            #     })

            # 添加消息到队列
            await add_to_queue({
                "process_message": process_message,
            })

            # 取出队列中的消息并发送
            while not MessageQueue.empty():
                message = await MessageQueue.get()
                await websocket.send_json(message)

            # 如果适应度降低则是不好的变异方向,则放弃,重新变异
            if best_fitness_float<=cur_max_fitness_float:
                continue
            else:
                # 适应度增加,则跳出循环
                break

        # 取出队列中的消息并发送
        # while not MessageQueue.empty():
        #     message = await MessageQueue.get()
        #     await websocket.send_json(message)
            
        process_message = '尝试找到进化方向完成'
        print(process_message)
        # await websocket.send_json({
        #     "process_message": process_message,
        #     })
        
        # 添加消息到队列
        await add_to_queue({
            "process_message": process_message,
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)

        # 变异方向是增加适应度的,则继续进化,基因继承和适应度继承
        BestSchedule = CurSchedule
        cur_max_fitness_float = best_fitness_float

        process_message = '处理种群信息可视化开始'
        print(process_message)
        # await websocket.send_json({
        #     "process_message": process_message,
        #     })
        # 打印当前代的最优适应度
        generation_int += 1
        # print(f"Generation {generation_int}: Best Fitness = {best_fitness_float}")
        # 获得当前代的排课信息
        # current_generation_info_dict = await asyncio.to_thread(CurSchedule.display)
        current_generation_info_dict = CurSchedule.display()
        # 获得已排课的课程数
        total_assigned_course_number = len(current_generation_info_dict['room_assigned_course_schedule_dict'].keys())
        # 字典添加代数和适应度
        current_generation_info_dict['generation_int']=generation_int
        # print('当前代数:',current_generation_info_dict['generation_int'])
        current_generation_info_dict['fitness_float']=best_fitness_float
        current_generation_info_dict['total_assigned_course_number']=total_assigned_course_number

        # time_used = public_function.get_time_used(start_time)
        process_message = '处理种群信息可视化完成'
        print(process_message)
        # await websocket.send_json({
        #     "process_message": process_message,
        #     })

        # 添加消息到队列
        await add_to_queue({
            "process_message": process_message,
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)

        # 存储全局课程排课信息
        all_generation_info_dict[generation_int] = current_generation_info_dict

        # 将内容发送给前端实时显示
        # await websocket.send_json({
        #     "generation": generation_int,"best_fitness":best_fitness_float,
        #     'current_generation_info_dict':current_generation_info_dict
        #     })
        # 添加消息到队列
        await add_to_queue({
            "generation": generation_int,"best_fitness":best_fitness_float,
            'current_generation_info_dict':current_generation_info_dict
        })

        # 取出队列中的消息并发送
        while not MessageQueue.empty():
            message = await MessageQueue.get()
            await websocket.send_json(message)

@router.get("/download_schedule_to_excel")
async def download_schedule_to_excel(generation: int = 0,fitness: float = 0.0):

    # 获取请求体中的 JSON 数据
    # request_data = await request.json()
    # cur_generation_int = request_data.get('generation_int')
    # 获取指定代数的排课信息
    cur_generation_info_dict = all_generation_info_dict.get(generation,None)
    if cur_generation_info_dict ==None:
        raise fastapi.HTTPException(status_code=404, detail="Schedule not found for this generation")

    # 创建一个工作簿
    Workbook = openpyxl.Workbook()
    ws = Workbook.active
    base_file_name_str = f"schedule_generation_{generation}_fitness_{fitness}"
    ws.title = base_file_name_str

    # 添加表头
    headers = ["上课周数","上课星期","上课时间","校区","上课教室","排课类别","课程名称", "授课教师id","课程id","教室id","课程总学时要求","已排总学时","本周排课次数","已排学时和要求学时的距离","教师当天上课讲数","教师本周上课讲数","课程优先级", "选课学生", "教师不可上课时间段","上课时间冲突学生", "是否处于教师不可上课时间"]
    ws.append(headers)

    # 填充排课数据
    for schedule in cur_generation_info_dict['room_assigned_course_schedule_dict'].values():
        row = [
            schedule['week'],
            schedule['day'],
            schedule['time'],
            schedule['campus_area_str'],

            schedule['room'],
            schedule['schedule_course_type'],
            schedule['course'],
            schedule['teacher'],

            schedule['course_id'],
            schedule['room_id'],
            
            schedule['cur_schedule_total_study_hour_float'],
            schedule['cur_study_hour'],
            schedule['schedule_week_count_int'],
            schedule['study_hour_distance_float'],

            schedule['course_teacher_day_study_hour_str'],
            schedule['course_teacher_week_study_hour_str'],

            schedule['priority'],
            schedule['enrolled_student'],
            schedule['unavailable_timeslots_teacher'],
            schedule['conflict'],

            schedule['availability'],

        ]
        ws.append(row)

    # 设置列宽自适应
    for col_num in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        max_length = 0
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

        # 设置列内文本水平左对齐,垂直居中
        for cell in ws[col_letter]:
            cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

    # 将工作簿保存到内存中的字节流
    file_stream = io.BytesIO()
    Workbook.save(file_stream)
    file_stream.seek(0)

    # 对文件名进行 URL 编码，避免特殊字符引发编码错误
    encoded_filename = urllib.parse.quote(base_file_name_str+'.xlsx')

    return fastapi.responses.StreamingResponse(
        file_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )
